import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def build_payroll_excel(lines: list[dict], period_id: int) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    # Header style
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["№", "Сотрудник", "Тип", "Рейсов", "Сумма рейсов", "Часов", "Сумма часов", "Итого", "Авансы", "К выплате"]
    ws.append(headers)

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for i, line in enumerate(lines, 1):
        emp_type = "Водитель" if line["employee_type"] == "driver" else "Оператор"
        advances = float(line.get("advances_amount", 0))
        total = float(line["total_amount"])
        correction = float(line.get("manual_correction", 0))
        payable = total + correction - advances
        row = [
            i,
            line.get("employee_name", ""),
            emp_type,
            line["trips_count"],
            line["trips_amount"],
            line["hours_total"],
            line["hours_amount"],
            total,
            advances,
            payable,
        ]
        ws.append(row)
        for col_idx in range(1, len(row) + 1):
            ws.cell(row=i + 1, column=col_idx).border = thin_border

    # Totals
    total_row = len(lines) + 2
    ws.cell(row=total_row, column=2, value="ИТОГО").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=sum(l["trips_amount"] for l in lines)).font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(l["hours_amount"] for l in lines)).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=sum(float(l["total_amount"]) for l in lines)).font = Font(bold=True)
    ws.cell(row=total_row, column=9, value=sum(float(l.get("advances_amount", 0)) for l in lines)).font = Font(bold=True)

    total_payable = sum(
        float(l["total_amount"]) + float(l.get("manual_correction", 0)) - float(l.get("advances_amount", 0))
        for l in lines
    )
    ws.cell(row=total_row, column=10, value=total_payable).font = Font(bold=True)

    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col_idx).border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
